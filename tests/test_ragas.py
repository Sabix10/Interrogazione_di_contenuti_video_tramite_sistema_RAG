import os
import json
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from ragas.metrics import faithfulness, answer_relevancy
from datasets import Dataset, Features, Sequence, Value

from app import (
    app,
    client,
    MODEL_NAME,
    SYSTEM_PROMPT,
    evaluate,
    RunConfig,
    create_db_from_youtube_video_url,
    get_similarity_from_query
)

def carica_dataset_multivideo(file_path="dataset_ragas_multi_video.json"):
    """Carica il dataset strutturato contente tutti i video e le rispettive domande."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File '{file_path}' non trovato nella directory.")
    with open(file_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def esporta_excel_multischermo(risultati_globali, file_output="ragas/Report_Finale_MultiVideo.xlsx"):
    """
    Crea un file Excel con una scheda dedicata
    a ciascun video elaborato e formule native per le medie matematiche.
    """
    os.makedirs("ragas", exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1B365D")
    KPI_LABEL_FONT = Font(name="Calibri", size=9, bold=True, color="555555")
    KPI_VALUE_FONT = Font(name="Calibri", size=14, bold=True, color="1B365D")
    DATA_FONT = Font(name="Calibri", size=11, color="000000")
    BOLD_FONT = Font(name="Calibri", size=11, bold=True)
    ZEBRA_FILL = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
    KPI_FILL = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    BORDER_STYLING = Border(
        left=Side(style='thin', color='D0D7DE'), right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'), bottom=Side(style='thin', color='D0D7DE')
    )

    for nome_video, dati in risultati_globali.items():
        ws = wb.create_sheet(title=nome_video[:31])
        ws.views.sheetView[0].showGridLines = True

        # Intestazione Scheda
        ws['A1'] = f"REPORT RAGAS - {nome_video}"
        ws['A1'].font = TITLE_FONT
        ws['A3'] = "Video URL:"
        ws['A3'].font = BOLD_FONT
        ws['B3'] = dati['url']
        ws['B3'].font = Font(name="Calibri", size=10, color="0033CC", underline="single")

        report_rows = dati['report']

        ws.merge_cells('D2:E2')
        ws['D2'] = "FAITHFULNESS MEDIA (PERTINENTI)"
        ws['D2'].font = KPI_LABEL_FONT
        ws['D2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D2'].fill = KPI_FILL
        ws.merge_cells('D3:E3')
        ws['D3'] = "=AVERAGE(E6:E8)"
        ws['D3'].font = KPI_VALUE_FONT
        ws['D3'].number_format = '0.0000'
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D3'].fill = KPI_FILL

        ws.merge_cells('F2:G2')
        ws['F2'] = "ANSWER RELEVANCY MEDIA (TOTALE)"
        ws['F2'].font = KPI_LABEL_FONT
        ws['F2'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F2'].fill = KPI_FILL
        ws.merge_cells('F3:G3')
        ws['F3'] = f"=AVERAGE(F6:F{len(report_rows) + 5})"
        ws['F3'].font = KPI_VALUE_FONT
        ws['F3'].number_format = '0.0000'
        ws['F3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F3'].fill = KPI_FILL

        for r in ws['D2:G3']:
            for c in r: c.border = BORDER_STYLING

        # Intestazioni tabella
        headers = ["ID", "CATEGORIA DOMANDA", "DOMANDA", "RISPOSTA CHATBOT (QWEN)", "FAITHFULNESS", "ANSWER RELEVANCY"]
        start_row = 5
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=text)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center' if col_idx in [1, 5, 6] else 'left', vertical='center')
            cell.border = BORDER_STYLING
        ws.row_dimensions[start_row].height = 26

        # Scrittura righe dati
        current_row = start_row + 1
        for item in report_rows:
            f_val = "N/A" if item['faithfulness'] is None or np.isnan(item['faithfulness']) else item['faithfulness']
            r_val = 0.0000 if item['answer_relevancy'] is None or np.isnan(item['answer_relevancy']) else item[
                'answer_relevancy']

            row_values = [item['id'], item['categoria'], item['domanda'], item['risposta'], f_val, r_val]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = BORDER_STYLING
                if current_row % 2 == 0: cell.fill = ZEBRA_FILL

                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx in [2, 3, 4]:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)
                elif col_idx in [5, 6]:
                    cell.alignment = Alignment(horizontal='center')
                    if isinstance(val, (int, float)): cell.number_format = '0.0000'
            ws.row_dimensions[current_row].height = 36
            current_row += 1

        # Riga di riepilogo
        ws.cell(row=current_row, column=2, value="MEDIA TOTALE").font = BOLD_FONT
        cell_f = ws.cell(row=current_row, column=5, value=f"=AVERAGE(E6:E{current_row - 1})")
        cell_f.font = BOLD_FONT
        cell_f.number_format = '0.0000'
        cell_f.alignment = Alignment(horizontal='center')

        cell_r = ws.cell(row=current_row, column=6, value=f"=AVERAGE(F6:F{current_row - 1})")
        cell_r.font = BOLD_FONT
        cell_r.number_format = '0.0000'
        cell_r.alignment = Alignment(horizontal='center')

        double_bottom = Border(bottom=Side(style='double', color='1B365D'), top=Side(style='thin', color='D0D7DE'))
        for c in range(1, 7): ws.cell(row=current_row, column=c).border = double_bottom

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 38
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18

    wb.save(file_output)
    print(f"\nReport Excel Globale salvato con successo in: {file_output}")


def run_ragas_evaluation(question, answer, contexts):
    """
    Calcola le metriche di valutazione RAGAS.
    """

    # Fallback
    if contexts is None:
        contexts = []

    data = {
        "question": [question],
        "answer": [answer],
        "contexts": [contexts]
    }

    schema_features = Features({
        "question": Value("string"),
        "answer": Value("string"),
        "contexts": Sequence(Value("string"))
    })

    dataset = Dataset.from_dict(data, features=schema_features)

    try:
        print("\n [RAGAS] Calcolo delle metriche di valutazione in corso...")

        # Configurazione del timeout interno di Ragas a 300 secondi per CPU locali
        configurazione_run = RunConfig(timeout=300)

        result = evaluate(
            dataset=dataset,
            metrics=[faithfulness, answer_relevancy],
            run_config=configurazione_run
        )

        # Estrazione diretta (i punteggi del dizionario result sono float singoli)
        faith_score = result['faithfulness']
        rel_score = result['answer_relevancy']

        print("=============================================")
        print("REPORT DI VALUTAZIONE RAGAS (Local Judge)")
        print(f"FAITHFULNESS: {faith_score:.4f}")
        print(f"ANSWER RELEVANCY: {rel_score:.4f}")
        print("=============================================\n")
        return result
    except Exception as e:
        import traceback
        print(f"Impossibile completare la valutazione RAGAS: {e}")
        traceback.print_exc()
        return None


def avvia_stress_test_multivideo(file_dataset="dataset_ragas_multi_video.json"):
    """
    Itera sui video, esegue il retrieval, interroga il LLM e delega
    la valutazione a run_ragas_evaluation prima di esportare in Excel.
    """
    with app.app_context():
        print(f"\nAVVIO STRESS TEST MULTI-VIDEO (Dataset: {file_dataset})")
        print("=" * 80)

        dataset = carica_dataset_multivideo(file_dataset)
        risultati_completi = {}

        for nome_video, info in dataset.items():
            url = info['url']
            domande_dizionario = info['domande']

            print(f"\n\n ==============================================================")

            try:
                create_db_from_youtube_video_url(url)
            except Exception as e:
                print(f"Impossibile caricare la trascrizione del video '{nome_video}': {e}")
                continue

            report_video = []
            contatore = 1

            for categoria, lista_domande in domande_dizionario.items():
                print(f"\nCategoria: [{categoria}]")
                print("   " + "-" * 55)

                for dom in lista_domande:
                    print(f"\n   [Test {contatore}/9] Domanda: '{dom}'")

                    # Retrieval da FAISS
                    retrieved_chunks = get_similarity_from_query(dom)

                    # Generazione risposta
                    if not retrieved_chunks:
                        bot_reply = "Non lo so."
                    else:
                        retrieved_context = "\n\n".join(retrieved_chunks)
                        augmented_prompt = (
                            f"Usa le informazioni contenute nel seguente contesto per rispondere alla domanda finale.\n\n"
                            f"### INIZIO CONTESTO ###\n"
                            f"{retrieved_context}\n"
                            f"### FINE CONTESTO ###\n\n"
                            f"DOMANDA: {dom}"
                        )
                        messages = [
                            {"role": "system", "content": SYSTEM_PROMPT},
                            {"role": "user", "content": augmented_prompt}
                        ]

                        try:
                            response = client.chat.completions.create(
                                model=MODEL_NAME, messages=messages, temperature=0.1, max_tokens=600
                            )
                            bot_reply = response.choices[0].message.content
                        except Exception as e:
                            print(f"      Errore LLM: {e}")
                            bot_reply = "ERRORE GENERAZIONE LLM"

                    print(f"      Risposta: {bot_reply}")

                    risultato_ragas = run_ragas_evaluation(dom, bot_reply, retrieved_chunks)

                    if risultato_ragas is not None:
                        s_faith = risultato_ragas['faithfulness']
                        s_relev = risultato_ragas['answer_relevancy']
                    else:
                        s_faith, s_relev = float('nan'), float('nan')

                    # Salvataggio nel dizionario dei report del video corrente
                    report_video.append({
                        "id": contatore,
                        "categoria": categoria,
                        "domanda": dom,
                        "risposta": bot_reply,
                        "faithfulness": s_faith,
                        "answer_relevancy": s_relev
                    })
                    contatore += 1

            risultati_completi[nome_video] = {
                "url": url,
                "report": report_video
            }

        # Generazione finale del file Excel multi-scheda
        esporta_excel_multischermo(risultati_completi)
        print("\nTUTTI I TEST MULTI-VIDEO SONO STATI COMPLETATI CON SUCCESSO!")

if __name__ == "__main__":
    avvia_stress_test_multivideo("dataset_ragas_multi_video.json")